"""
Импорт результатов экспертной валидации (P2-1.3, D3.5–D3.7).

ОСНОВНОЙ режим (пакет 3 юристов) — расчёт метрик по возврату панели:
  uv run python scripts/import_expert_validation.py \\
    --files data/expert_validation/expert_1_annotated.xlsx \\
            data/expert_validation/expert_2_annotated.xlsx \\
            data/expert_validation/expert_3_annotated.xlsx \\
    --keymap data/expert_validation/keymap.json \\
    [--llm-results data/llm_validation_24.json]

Метрики:
  • Fleiss κ по задачам (5 критериев + агрегированно) и по рёбрам (D3.6)
  • Cohen κ: LLM-валидатор vs консенсус юристов (D3.5/3c) — если задан --llm-results
  • precision рёбер с Wilson CI (D3.7), отдельно доктринальные / структурные
  • доля альтернативных цепочек по типам задач (D3.5)

LEGACY режим (одиночный файл Фазы 1, 3 критерия) — флаг --legacy.

Парные функции чистые (берут разобранные аннотации) → юнит-тестируемы на синтетике.
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from mhgb.analysis.statistics import compute_fleiss_kappa, wilson_ci, _cohens_kappa

# 5 критериев задач (порядок = шкала LLM-валидатора)
CRIT_KEYS = ["К1", "К2", "К3", "К4", "К5"]
# соответствие критериев проверкам LLM-валидатора (для Cohen κ, D3.5/3c)
CRIT_TO_LLM = {
    "К1": "gold_chain_matches_norm_ids",
    "К2": "task_solvable_from_context",
    "К3": "fabula_complete",
    "К4": "no_logical_gaps",
    "К5": "answer_consistent",
}
ALT_NONTRIVIAL = {"незначительная", "существенная"}


def _to_bool(val) -> bool | None:
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("да", "yes", "true", "1", "+"):
        return True
    if s in ("нет", "no", "false", "0", "-"):
        return False
    return None


# ===========================================================================
# ПАКЕТ 3 ЮРИСТОВ (P2-1.3) — парсинг + метрики
# ===========================================================================

def load_package_annotation(xlsx_path: str) -> dict:
    """Читает один заполненный пакет (листы «Задачи» + «Рёбра»).

    Возвращает {"tasks": {anon_id: {К1..К5: bool|None, alt: str|None, comment}},
                "edges": {anon_id: {correct: bool|None, comment}}}.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    def _col_index(ws, predicate):
        headers = [c.value for c in ws[1]]
        return {i: h for i, h in enumerate(headers) if h and predicate(str(h))}

    # --- Задачи ---
    wt = wb["Задачи"]
    headers = [c.value for c in wt[1]]
    id_col = headers.index("ID")
    crit_cols = {}
    for key in CRIT_KEYS:
        for i, h in enumerate(headers):
            if h and str(h).startswith(key):
                crit_cols[key] = i
                break
    alt_col = next((i for i, h in enumerate(headers) if h and "льтернатив" in str(h)), None)
    com_col = next((i for i, h in enumerate(headers) if h and str(h).strip() == "Комментарий"), None)

    tasks: dict[str, dict] = {}
    for row in wt.iter_rows(min_row=2, values_only=True):
        anon = row[id_col]
        if not anon:
            continue
        rec = {k: _to_bool(row[crit_cols[k]]) for k in CRIT_KEYS if k in crit_cols}
        rec["alt"] = (str(row[alt_col]).strip().lower() if alt_col is not None and row[alt_col] else None)
        rec["comment"] = (str(row[com_col]) if com_col is not None and row[com_col] else "")
        tasks[str(anon)] = rec

    # --- Рёбра ---
    we = wb["Рёбра"]
    eh = [c.value for c in we[1]]
    eid_col = eh.index("ID")
    ok_col = next(i for i, h in enumerate(eh) if h and "корректна" in str(h))
    ecom_col = next((i for i, h in enumerate(eh) if h and str(h).strip() == "Комментарий"), None)

    edges: dict[str, dict] = {}
    for row in we.iter_rows(min_row=2, values_only=True):
        anon = row[eid_col]
        if not anon:
            continue
        edges[str(anon)] = {
            "correct": _to_bool(row[ok_col]),
            "comment": (str(row[ecom_col]) if ecom_col is not None and row[ecom_col] else ""),
        }
    return {"tasks": tasks, "edges": edges}


def _matrix_for_criterion(annots: list[dict], key: str) -> list[list[int]]:
    """N задач × R юристов матрица меток (0/1) по одному критерию; строки с пропусками отбрасываются."""
    ids = sorted(annots[0]["tasks"])
    matrix = []
    for tid in ids:
        labels = [a["tasks"].get(tid, {}).get(key) for a in annots]
        if any(v is None for v in labels):
            continue
        matrix.append([int(v) for v in labels])
    return matrix


def compute_task_fleiss(annots: list[dict]) -> dict:
    """Fleiss κ по каждому критерию + агрегированный (пул всех задача×критерий)."""
    per_crit, pooled = {}, []
    for key in CRIT_KEYS:
        m = _matrix_for_criterion(annots, key)
        per_crit[key] = compute_fleiss_kappa(m) if len(m) >= 1 else None
        pooled.extend(m)
    aggregate = compute_fleiss_kappa(pooled) if pooled else None
    return {"per_criterion": per_crit, "aggregate": aggregate, "n_items_pooled": len(pooled)}


def compute_edge_fleiss(annots: list[dict]) -> dict:
    ids = sorted(annots[0]["edges"])
    matrix = []
    for eid in ids:
        labels = [a["edges"].get(eid, {}).get("correct") for a in annots]
        if any(v is None for v in labels):
            continue
        matrix.append([int(v) for v in labels])
    return {"kappa": compute_fleiss_kappa(matrix) if matrix else None, "n_items": len(matrix)}


def compute_edge_precision(annots: list[dict], keymap: dict) -> dict:
    """Majority-vote precision рёбер + Wilson CI, отдельно по группам (доктринальные/структурные)."""
    groups: dict[str, list[int]] = defaultdict(list)  # group → [1 если majority=корректно]
    ids = sorted(annots[0]["edges"])
    for eid in ids:
        labels = [a["edges"].get(eid, {}).get("correct") for a in annots]
        labels = [v for v in labels if v is not None]
        if not labels:
            continue
        majority = sum(labels) > len(labels) / 2
        group = keymap.get("edges", {}).get(eid, {}).get("group", "unknown")
        groups[group].append(int(majority))
        groups["all"].append(int(majority))

    out = {}
    for g, vals in groups.items():
        n, s = len(vals), sum(vals)
        lo, hi = wilson_ci(s, n) if n else (None, None)
        out[g] = {"precision": s / n if n else None, "n": n, "correct": s,
                  "wilson_ci": [lo, hi]}
    return out


def compute_alternative_fraction(annots: list[dict], keymap: dict, type_by_id: dict) -> dict:
    """Доля задач с нетривиальной альтернативой (majority), по типам задач (D3.5)."""
    by_type: dict[str, list[int]] = defaultdict(list)
    ids = sorted(annots[0]["tasks"])
    for tid in ids:
        votes = [a["tasks"].get(tid, {}).get("alt") for a in annots]
        votes = [v for v in votes if v]
        if not votes:
            continue
        # нетривиально, если большинство отметили незначительную/существенную
        nontrivial = sum(1 for v in votes if v in ALT_NONTRIVIAL) > len(votes) / 2
        real = keymap.get("tasks", {}).get(tid)
        ttype = type_by_id.get(real, "unknown")
        by_type[ttype].append(int(nontrivial))
        by_type["all"].append(int(nontrivial))
    return {t: {"fraction": sum(v) / len(v), "n": len(v), "nontrivial": sum(v)}
            for t, v in by_type.items()}


def compute_llm_vs_expert_cohen(annots: list[dict], keymap: dict, llm_results: dict) -> dict:
    """Cohen κ: LLM-валидатор vs консенсус юристов, по критериям + агрегированно (D3.5/3c).

    llm_results: {real_task_id: {checks: {llm_check_name: bool}}} (или {real_task_id: {llm_check_name: bool}}).
    """
    def _llm_check(real_id: str, llm_name: str):
        rec = llm_results.get(real_id) or llm_results.get(_base(real_id))
        if rec is None:
            return None
        checks = rec.get("checks", rec)
        return checks.get(llm_name)

    ids = sorted(annots[0]["tasks"])
    per_crit, pooled_e, pooled_l = {}, [], []
    for key in CRIT_KEYS:
        exp_lbls, llm_lbls = [], []
        for tid in ids:
            votes = [a["tasks"].get(tid, {}).get(key) for a in annots]
            votes = [v for v in votes if v is not None]
            if not votes:
                continue
            real = keymap.get("tasks", {}).get(tid)
            lv = _llm_check(real, CRIT_TO_LLM[key])
            if lv is None:
                continue
            consensus = int(sum(votes) > len(votes) / 2)
            exp_lbls.append(consensus)
            llm_lbls.append(int(bool(lv)))
        per_crit[key] = _cohens_kappa(exp_lbls, llm_lbls) if len(exp_lbls) >= 2 else None
        pooled_e.extend(exp_lbls)
        pooled_l.extend(llm_lbls)
    agg = _cohens_kappa(pooled_e, pooled_l) if len(pooled_e) >= 2 else None
    return {"per_criterion": per_crit, "aggregate": agg, "n_pooled": len(pooled_e)}


def compute_llm_vs_expert_cohen_edges(annots: list[dict], keymap: dict, llm_edge: dict) -> dict:
    """Cohen κ: LLM (llama) vs консенсус юристов по РЁБРАМ (бинарно да/нет).

    llm_edge: {E-id: {"correct": bool}} (из run_llm_edge_validation.py).
    Возвращает κ общий + отдельно doctrinal / structural (как в precision).
    Консенсус юриста по ребру = majority из 3 (да/нет).
    """
    ids = sorted(annots[0]["edges"])
    groups: dict[str, tuple[list, list]] = {"all": ([], []), "doctrinal": ([], []), "structural": ([], [])}
    for eid in ids:
        votes = [a["edges"].get(eid, {}).get("correct") for a in annots]
        votes = [v for v in votes if v is not None]
        if not votes:
            continue
        lv = llm_edge.get(eid, {}).get("correct")
        if lv is None:
            continue
        consensus = int(sum(votes) > len(votes) / 2)
        g = keymap.get("edges", {}).get(eid, {}).get("group", "unknown")
        for key in ("all", g):
            if key in groups:
                groups[key][0].append(consensus)
                groups[key][1].append(int(bool(lv)))
    out = {}
    for key, (e_lbls, l_lbls) in groups.items():
        out[key] = {"kappa": _cohens_kappa(e_lbls, l_lbls) if len(e_lbls) >= 2 else None,
                    "n": len(e_lbls)}
    return out


def _base(tid: str) -> str:
    return tid.replace("_open", "").replace("_closed", "")


def run_package_report(files: list[str], keymap_path: str,
                       tasks_path: str = "data/expert_subset_24.jsonl",
                       llm_results_path: str | None = None,
                       llm_edge_results_path: str | None = None) -> dict:
    annots = [load_package_annotation(f) for f in files]
    keymap = json.loads(Path(keymap_path).read_text(encoding="utf-8"))

    type_by_id = {}
    tp = Path(tasks_path)
    if tp.exists():
        for line in tp.read_text(encoding="utf-8").splitlines():
            if line.strip():
                t = json.loads(line)
                type_by_id[t["id"]] = t.get("type", "unknown")

    report = {
        "n_experts": len(annots),
        "task_fleiss": compute_task_fleiss(annots),
        "edge_fleiss": compute_edge_fleiss(annots),
        "edge_precision": compute_edge_precision(annots, keymap),
        "alternative_fraction": compute_alternative_fraction(annots, keymap, type_by_id),
    }
    if llm_results_path and Path(llm_results_path).exists():
        llm = json.loads(Path(llm_results_path).read_text(encoding="utf-8"))
        report["llm_vs_expert_cohen"] = compute_llm_vs_expert_cohen(annots, keymap, llm)
    if llm_edge_results_path and Path(llm_edge_results_path).exists():
        llm_e = json.loads(Path(llm_edge_results_path).read_text(encoding="utf-8"))
        report["llm_vs_expert_cohen_edges"] = compute_llm_vs_expert_cohen_edges(annots, keymap, llm_e)
    return report


def _print_package_report(r: dict) -> None:
    print(f"\n=== Экспертная валидация: {r['n_experts']} юристов ===\n")
    tf = r["task_fleiss"]
    print("Fleiss κ по задачам (5 критериев):")
    for k, v in tf["per_criterion"].items():
        print(f"  {k}: {v:.3f}" if v is not None else f"  {k}: —")
    print(f"  агрегированный: {tf['aggregate']:.3f}" if tf["aggregate"] is not None else "  агрегированный: —")
    ef = r["edge_fleiss"]
    print(f"\nFleiss κ по рёбрам: {ef['kappa']:.3f} (n={ef['n_items']})" if ef["kappa"] is not None else "\nFleiss κ по рёбрам: —")
    print("\nPrecision рёбер (majority + Wilson 95% CI):")
    for g, d in r["edge_precision"].items():
        if d["precision"] is None:
            continue
        lo, hi = d["wilson_ci"]
        print(f"  {g:11}: {d['precision']:.3f} [{lo:.3f}, {hi:.3f}]  (n={d['n']}, корр.={d['correct']})")
    print("\nДоля альтернативных цепочек (нетривиальные, по типам):")
    for t, d in r["alternative_fraction"].items():
        print(f"  {t:22}: {d['fraction']:.3f}  (n={d['n']}, нетривиальных={d['nontrivial']})")
    if "llm_vs_expert_cohen" in r:
        c = r["llm_vs_expert_cohen"]
        print("\nCohen κ ЗАДАЧИ: LLM-валидатор (llama) vs консенсус юристов:")
        for k, v in c["per_criterion"].items():
            print(f"  {k}: {v:.3f}" if v is not None else f"  {k}: —")
        print(f"  агрегированный: {c['aggregate']:.3f} (n_pooled={c['n_pooled']})"
              if c["aggregate"] is not None else "  агрегированный: —")
    if "llm_vs_expert_cohen_edges" in r:
        ce = r["llm_vs_expert_cohen_edges"]
        print("\nCohen κ РЁБРА: LLM (llama) vs консенсус юристов:")
        for g in ("all", "doctrinal", "structural"):
            d = ce.get(g)
            if d and d["kappa"] is not None:
                print(f"  {g:11}: {d['kappa']:.3f} (n={d['n']})")


# ---------------------------------------------------------------------------
# LEGACY (Фаза 1): одиночный файл, 3 критерия
# ---------------------------------------------------------------------------

def load_expert_results(xlsx_path: str) -> list[dict]:
    """
    Читает Excel и возвращает список дiktов с результатами валидации.

    Каждый dict содержит:
      task_id, fabula_ok, chain_ok, answer_ok, is_valid, notes
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Задачи"]

    headers = [cell.value for cell in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h}

    required = {"task_id", "expert_fabula_ok", "expert_chain_ok", "expert_answer_ok"}
    missing = required - set(idx.keys())
    if missing:
        raise ValueError(f"В Excel отсутствуют столбцы: {missing}")

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        task_id = row[idx["task_id"]]
        if not task_id:
            continue

        def _to_bool(val) -> bool | None:
            if val is None:
                return None
            s = str(val).strip().lower()
            if s in ("да", "yes", "true", "1", "+"):
                return True
            if s in ("нет", "no", "false", "0", "-"):
                return False
            return None

        fabula_ok = _to_bool(row[idx["expert_fabula_ok"]])
        chain_ok  = _to_bool(row[idx["expert_chain_ok"]])
        answer_ok = _to_bool(row[idx["expert_answer_ok"]])
        notes     = row[idx.get("expert_notes", -1)] if "expert_notes" in idx else ""

        if fabula_ok is None or chain_ok is None or answer_ok is None:
            continue  # строка не заполнена

        results.append({
            "task_id":   str(task_id),
            "fabula_ok": fabula_ok,
            "chain_ok":  chain_ok,
            "answer_ok": answer_ok,
            "is_valid":  fabula_ok and chain_ok and answer_ok,
            "notes":     str(notes) if notes else "",
        })

    return results


# ---------------------------------------------------------------------------
# Сохранение в MongoDB
# ---------------------------------------------------------------------------

def save_to_mongo(
    results: list[dict],
    reviewer: str,
    batch_id: str,
) -> None:
    """Сохраняет результаты в коллекцию validation_logs."""
    from mhgb.storage.mongo_client import MongoStorage
    from mhgb.storage.schemas import ValidationLog

    storage = MongoStorage()

    for r in results:
        checks = {
            "fabula_complete":  r["fabula_ok"],
            "no_logical_gaps":  r["chain_ok"],
            "answer_consistent": r["answer_ok"],
        }
        log = ValidationLog(
            task_id=r["task_id"],
            task_batch_id=batch_id,
            validator="expert",
            is_valid=r["is_valid"],
            checks=checks,
            issues=[f"[{reviewer}] {r['notes']}"] if r["notes"] else [],
            judge_explanation=f"Рецензент: {reviewer}",
        )
        storage.db["validation_logs"].insert_one(log.model_dump())

    print(f"Сохранено {len(results)} результатов в MongoDB (reviewer={reviewer})")


# ---------------------------------------------------------------------------
# Сравнение двух рецензентов
# ---------------------------------------------------------------------------

def compare_two_reviewers(file_a: str, file_b: str) -> None:
    """
    Загружает два Excel с результатами и вычисляет метрики согласия.
    """
    from mhgb.validation.expert_validation import compare_llm_vs_expert

    results_a = load_expert_results(file_a)
    results_b = load_expert_results(file_b)

    comparison = compare_llm_vs_expert(results_a, results_b)

    print(f"\n=== Сравнение двух рецензентов ===")
    print(f"Совпадающих задач: {comparison['total']}")
    print(f"Процент согласия:  {comparison['agreement']:.1%}")
    print(f"Cohen's κ:         {comparison['cohens_kappa']:.3f}")
    print(f"Оба валидны:       {comparison['counts']['both_valid']}")
    print(f"Оба отклонили:     {comparison['counts']['both_rejected']}")
    print(f"A валид, B откл.:  {comparison['counts']['llm_valid_expert_rejected']}")
    print(f"A откл., B валид.: {comparison['counts']['llm_rejected_expert_valid']}")

    if comparison["agreement"] >= 0.85:
        print("\n✅ Согласие ≥85% — LLM-валидатор можно масштабировать на весь датасет.")
    else:
        print(f"\n⚠️  Согласие {comparison['agreement']:.1%} < 85% — требуется дополнительный анализ.")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Импорт результатов экспертной валидации (P2-1.3)")
    # пакетный режим (по умолчанию)
    parser.add_argument("--files", nargs="+", help="заполненные пакеты 3 юристов (xlsx)")
    parser.add_argument("--keymap", help="keymap.json (anon→real)")
    parser.add_argument("--tasks", default="data/expert_subset_24.jsonl")
    parser.add_argument("--llm-results", help="JSON LLM-валидатора задач (Cohen κ задачи)")
    parser.add_argument("--llm-edge-results", help="JSON LLM-валидатора рёбер (Cohen κ рёбра)")
    parser.add_argument("--out", help="куда сохранить JSON-отчёт")
    # legacy
    parser.add_argument("--legacy", action="store_true", help="старый режим (1 файл, 3 критерия)")
    parser.add_argument("--input",     help="(legacy) путь к Excel")
    parser.add_argument("--reviewer",  default="Expert1", help="(legacy)")
    parser.add_argument("--batch-id",  default="", help="(legacy)")
    parser.add_argument("--save-mongo", action="store_true", help="(legacy)")
    parser.add_argument("--compare",   help="(legacy) второй Excel")
    args = parser.parse_args()

    # --- пакетный режим ---
    if not args.legacy:
        if not args.files or not args.keymap:
            parser.error("пакетный режим требует --files и --keymap (или используйте --legacy)")
        report = run_package_report(args.files, args.keymap, args.tasks,
                                    args.llm_results, args.llm_edge_results)
        _print_package_report(report)
        out = Path(args.out) if args.out else Path(args.files[0]).with_name("expert_validation_report.json")
        out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"\nОтчёт → {out}")
        return

    # --- legacy ---
    if not args.input:
        parser.error("legacy режим требует --input")
    results = load_expert_results(args.input)

    valid   = sum(1 for r in results if r["is_valid"])
    invalid = len(results) - valid
    print(f"Загружено {len(results)} результатов от {args.reviewer}")
    print(f"  Валидных:    {valid}")
    print(f"  Отклонённых: {invalid}")
    if results:
        print(f"  Процент валидных: {valid / len(results):.1%}")

    if args.save_mongo:
        save_to_mongo(results, args.reviewer, args.batch_id)

    if args.compare:
        compare_two_reviewers(args.input, args.compare)

    # Сохраняем JSON-отчёт рядом с Excel
    report_path = Path(args.input).with_suffix(".report.json")
    report = {
        "reviewer": args.reviewer,
        "total": len(results),
        "valid_count": valid,
        "rejected_count": invalid,
        "valid_rate": valid / len(results) if results else 0.0,
        "results": results,
    }
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"Отчёт сохранён → {report_path}")


if __name__ == "__main__":
    main()
