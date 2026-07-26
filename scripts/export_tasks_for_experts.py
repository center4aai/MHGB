"""Единый Excel-пакет для валидации юристами (P2-1.2, D3.3/D3.4/D3.5).

Один файл, 3 подблока (листа), одинаковый для всех 3 юристов (полное перекрытие → Fleiss κ):
  1. «Задачи»     — 24 задачи (`expert_subset_24.jsonl`): 5 критериев (да/нет, в шкале
                    LLM-валидатора → Cohen κ) + комментарий + вопрос об альтернативной
                    цепочке (D3.5: нет / незначительная / существенная).
  2. «Рёбра»      — 70 рёбер (`edge_validation_sample.jsonl`): корректно/некорректно + коммент.
  3. «Инструкция» — шкалы, 3 подблока, пограничные примеры.

Анонимные ID (T01…T24, E01…E70); карта anon→real → `keymap.json` (для импорта P2-1.3,
юристам НЕ отдаётся). Реальные task_id / source-target / LLM-explanation юристы не видят.

Запуск:
  uv run python scripts/export_tasks_for_experts.py            # пакет + 3 копии + keymap
  uv run python scripts/export_tasks_for_experts.py --copies 0 # только template + keymap
  uv run python scripts/export_tasks_for_experts.py --legacy   # старый Фаза-1 экспорт
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

ROOT = Path(__file__).resolve().parent.parent
TASKS_SUBSET = ROOT / "data" / "expert_subset_24.jsonl"
EDGES_SUBSET = ROOT / "data" / "edge_validation_sample.jsonl"
OUT_DIR = ROOT / "data" / "expert_validation"

TYPE_RU = {
    "issue_spotting":      "Issue Spotting",
    "rule_selection":      "Rule Selection",
    "conflict_resolution": "Conflict Resolution",
    "temporal_validity":   "Temporal Validity",
}
HOP_RU = {"shallow": "shallow (1–2)", "medium": "medium (3–4)", "deep": "deep (5+)"}

# 5 критериев — дословно из mhgb.validation.llm_validator (для сопоставимости Cohen κ)
TASK_CRITERIA = [
    ("К1. gold_chain ⊆ norm_ids",
     "Каждая норма из цепочки рассуждений присутствует в списке норм задачи."),
    ("К2. Задача решаема",
     "Вопрос однозначен; для решения достаточно фабулы и нормативной базы."),
    ("К3. Фабула полна",
     "Фабула содержит все необходимые факты (кто, что, когда). Для temporal — дата события."),
    ("К4. Нет логических пробелов",
     "Каждый шаг цепочки логически следует из предыдущего, без скачков и пропущенных норм."),
    ("К5. Ответ согласован",
     "Эталонный вывод логически следует из цепочки рассуждений."),
]
ALT_CHAIN_HEADER = "Альтернативная цепочка?"
ALT_CHAIN_VALUES = "нет,незначительная,существенная"
YES_NO = "да,нет"


def _read_jsonl(path: Path) -> list[dict]:
    with path.open(encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


def _gold_chain_pretty(gc: list[dict]) -> str:
    out = []
    for s in gc:
        if s.get("norm_id"):
            out.append(f"[шаг {s.get('step')}] {s['norm_id']}: {s.get('reasoning', '')}")
        else:
            out.append(f"[вывод] {s.get('conclusion', s.get('reasoning', ''))}")
    return "\n".join(out)


# ---------------------------------------------------------------------------
# Сборка пакета (Фаза 2)
# ---------------------------------------------------------------------------

def build_expert_package(out_path: Path, copies: int = 3) -> dict:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    tasks = _read_jsonl(TASKS_SUBSET)
    edges = _read_jsonl(EDGES_SUBSET)

    keymap = {"tasks": {}, "edges": {}}

    wb = openpyxl.Workbook()
    head_fill = PatternFill("solid", fgColor="4472C4")
    ask_fill = PatternFill("solid", fgColor="E2EFDA")
    head_font = Font(bold=True, color="FFFFFF")
    ask_font = Font(bold=True, color="2E5B2E")
    wrap = Alignment(wrap_text=True, vertical="top")

    # ---- Лист 1: Задачи ----
    ws = wb.active
    ws.title = "Задачи"
    crit_titles = [c[0] for c in TASK_CRITERIA]
    headers = ["ID", "Тип", "Глубина", "Фабула", "Вопрос", "Эталонный ответ",
               "Нормы", "Цепочка рассуждений", *crit_titles, "Комментарий", ALT_CHAIN_HEADER]
    n_meta = 8  # сколько колонок до критериев (A..H)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.alignment = wrap
        if col > n_meta:
            cell.fill, cell.font = ask_fill, ask_font
        else:
            cell.fill, cell.font = head_fill, head_font

    for i, t in enumerate(tasks, start=1):
        anon = f"T{i:02d}"
        keymap["tasks"][anon] = t["id"]
        norm_ids = [s["norm_id"] for s in t.get("gold_chain", []) if s.get("norm_id")]
        row = [
            anon,
            TYPE_RU.get(t.get("type", ""), t.get("type", "")),
            HOP_RU.get(t.get("hop_group", ""), t.get("hop_group", "")),
            t.get("fabula", ""),
            t.get("question", ""),
            t.get("answer", ""),
            ", ".join(dict.fromkeys(norm_ids)),
            _gold_chain_pretty(t.get("gold_chain", [])),
        ] + [""] * (len(TASK_CRITERIA) + 2)  # 5 критериев + коммент + альтернатива
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=i + 1, column=col, value=val)
            c.alignment = wrap

    n_task_rows = len(tasks) + 1
    # дропдауны: 5 критериев да/нет
    for j in range(len(TASK_CRITERIA)):
        col = get_column_letter(n_meta + 1 + j)
        dv = DataValidation(type="list", formula1=f'"{YES_NO}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{n_task_rows}")
    # альтернатива (последняя колонка)
    alt_col = get_column_letter(len(headers))
    dv_alt = DataValidation(type="list", formula1=f'"{ALT_CHAIN_VALUES}"', allow_blank=True)
    ws.add_data_validation(dv_alt)
    dv_alt.add(f"{alt_col}2:{alt_col}{n_task_rows}")

    widths = [6, 16, 12, 55, 40, 45, 22, 60] + [13] * len(TASK_CRITERIA) + [30, 20]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "I2"  # видны A–H при скролле к критериям
    ws.row_dimensions[1].height = 30

    # ---- Лист 2: Рёбра ----
    we = wb.create_sheet("Рёбра")
    e_headers = ["ID", "Тип связи", "Норма A (id)", "Норма A — текст",
                 "Норма B (id)", "Норма B — текст", "Связь корректна?", "Комментарий"]
    e_meta = 6
    for col, h in enumerate(e_headers, start=1):
        cell = we.cell(row=1, column=col, value=h)
        cell.alignment = wrap
        if col > e_meta:
            cell.fill, cell.font = ask_fill, ask_font
        else:
            cell.fill, cell.font = head_fill, head_font

    for i, e in enumerate(edges, start=1):
        anon = f"E{i:02d}"
        keymap["edges"][anon] = {"source": e["source"], "target": e["target"],
                                 "edge_type": e["edge_type"], "group": e["group"]}
        # тексты с заголовком; реальные id норм юрист видит (это не «ответ»),
        # но НЕ видит LLM-explanation (чтобы не якорить оценку) и не знает «правильного» вердикта
        src_txt = (f"{e['source_title']}\n{e['source_text']}").strip()
        tgt_txt = (f"{e['target_title']}\n{e['target_text']}").strip()
        row = [anon, e["edge_type"], e["source"], src_txt, e["target"], tgt_txt, "", ""]
        for col, val in enumerate(row, start=1):
            c = we.cell(row=i + 1, column=col, value=val)
            c.alignment = wrap

    n_edge_rows = len(edges) + 1
    ok_col = get_column_letter(e_meta + 1)
    dv_e = DataValidation(type="list", formula1=f'"{YES_NO}"', allow_blank=True)
    we.add_data_validation(dv_e)
    dv_e.add(f"{ok_col}2:{ok_col}{n_edge_rows}")
    e_widths = [6, 16, 14, 55, 14, 55, 16, 30]
    for col, w in enumerate(e_widths, start=1):
        we.column_dimensions[get_column_letter(col)].width = w
    we.freeze_panes = "G2"
    we.row_dimensions[1].height = 30

    # ---- Лист 3: Инструкция ----
    wi = wb.create_sheet("Инструкция")
    _fill_instruction(wi)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    # копии для 3 юристов (идентичные, полное перекрытие D3.3)
    copy_paths = []
    for k in range(1, copies + 1):
        cp = out_path.with_name(f"expert_{k}.xlsx")
        wb.save(cp)
        copy_paths.append(cp)

    keymap_path = OUT_DIR / "keymap.json"
    keymap_path.write_text(json.dumps(keymap, ensure_ascii=False, indent=2), encoding="utf-8")

    return {
        "n_tasks": len(tasks),
        "n_edges": len(edges),
        "template": out_path,
        "copies": copy_paths,
        "keymap": keymap_path,
    }


def _fill_instruction(ws) -> None:
    from openpyxl.styles import Font
    rows = [
        ["Инструкция для эксперта-юриста — пакет валидации MHGB"],
        [],
        ["Пакет состоит из двух рабочих листов. Заполняются только зелёные столбцы."],
        ["Все 3 эксперта получают ОДИНАКОВЫЙ набор (нужно для расчёта согласия)."],
        [],
        ["── Лист «Задачи» (24 шт): оцените 5 критериев + вопрос об альтернативе ──"],
        ["Критерий", "Что означает «да»", "Значения"],
        *[[t, d, "да / нет"] for t, d in TASK_CRITERIA],
        [ALT_CHAIN_HEADER,
         "Есть ли иная юридически валидная цепочка к тому же выводу? "
         "«незначительная» — мелкая вариация; «существенная» — принципиально другой путь.",
         "нет / незначительная / существенная"],
        ["Комментарий", "Замечания по любому критерию (необязательно).", "текст"],
        [],
        ["── Лист «Рёбра» (70 шт): корректна ли заявленная связь между нормами ──"],
        ["Связь корректна?",
         "Действительно ли между Нормой A и Нормой B есть указанное отношение "
         "(исключает / дополняет / приоритет / ссылается_на / применяется_к)? "
         "Направление НЕ оценивается — только наличие отношения.",
         "да / нет"],
        ["Комментарий", "Если «нет» — кратко почему (необязательно).", "текст"],
        [],
        ["── Пограничные примеры ──"],
        ["• Фабула без даты в temporal_validity → К3 = нет."],
        ["• Цепочка ссылается на норму, которой нет в списке норм → К1 = нет."],
        ["• Вывод верный, но один шаг логически не следует → К4 = нет, К5 может быть да."],
        ["• Связь «дополняет» между явно несвязанными статьями → Связь корректна = нет."],
        [],
        ["Если сомневаетесь — ставьте «нет» и опишите в комментарии."],
        ["По вопросам: polukoshko.marina@gmail.com"],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 75
    ws.column_dimensions["C"].width = 30


# ---------------------------------------------------------------------------
# Legacy Фаза-1 экспорт (сохранён для совместимости)
# ---------------------------------------------------------------------------

def export_to_excel(tasks_path: str, output_path: str, n_per_cell: int = 3, seed: int = 42) -> int:
    """Старый экспорт (3 критерия, пере-сэмплинг). Оставлен для обратной совместимости."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from mhgb.validation.expert_validation import select_stratified_sample

    tasks = _read_jsonl(Path(tasks_path))
    sample = select_stratified_sample(tasks, n_per_cell=n_per_cell, seed=seed)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Задачи"
    wrap = Alignment(wrap_text=True, vertical="top")
    headers = ["task_id", "Тип задачи", "Режим", "Сложность", "Фабула", "Вопрос",
               "Эталонный ответ", "expert_fabula_ok", "expert_chain_ok",
               "expert_answer_ok", "expert_notes"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)
    for r, t in enumerate(sample, start=2):
        vals = [t.get("id", ""), t.get("type", ""), t.get("mode", ""),
                t.get("hop_group", ""), t.get("fabula", ""), t.get("question", ""),
                t.get("answer", ""), "", "", "", ""]
        for col, v in enumerate(vals, start=1):
            ws.cell(row=r, column=col, value=v).alignment = wrap
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return len(sample)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="Экспорт пакета валидации для юристов (P2-1.2)")
    ap.add_argument("--out", default=str(OUT_DIR / "expert_package.xlsx"))
    ap.add_argument("--copies", type=int, default=3, help="сколько идентичных копий юристам")
    ap.add_argument("--legacy", action="store_true", help="старый Фаза-1 экспорт")
    ap.add_argument("--tasks", default="data/tasks_raw.jsonl", help="(legacy)")
    ap.add_argument("--n-per-cell", type=int, default=3, help="(legacy)")
    ap.add_argument("--seed", type=int, default=42, help="(legacy)")
    args = ap.parse_args()

    if args.legacy:
        n = export_to_excel(args.tasks, args.out, args.n_per_cell, args.seed)
        print(f"[legacy] экспортировано {n} задач → {args.out}")
        return

    info = build_expert_package(Path(args.out), copies=args.copies)
    print(f"✅ Пакет: {info['n_tasks']} задач + {info['n_edges']} рёбер")
    print(f"   template → {info['template'].relative_to(ROOT)}")
    for cp in info["copies"]:
        print(f"   копия    → {cp.relative_to(ROOT)}")
    print(f"   keymap   → {info['keymap'].relative_to(ROOT)}  (юристам НЕ отдавать)")


if __name__ == "__main__":
    main()
